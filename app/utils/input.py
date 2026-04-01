from pathlib import Path

import openpyxl

from app.core.models import Transaction

FieldMapping: dict[str, str] = {
    "Date": "date",
    "Platform": "platform",
    "Type": "type",
    "Symbol": "symbol",
    "Quantity": "quantity",
    "T. Price": "trade_price",
    "Amount": "amount",
    "Fees & Comm": "fees",
    "Tax Withholding": "tax_withholding",
    "Net Amount": "net_amount",
}


def read_transactions(file_path: Path) -> list[Transaction]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    assert ws

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    transactions: list[Transaction] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw = dict(zip(headers, row))
        mapped = {FieldMapping[k]: v for k, v in raw.items()}
        transactions.append(Transaction(**mapped))

    return transactions
